import os
import requests
import json
import random
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import csv

# Define la URL del endpoint del servidor
url = 'http://127.0.0.1:8000/api/config/'  # Cambia esta URL por la del endpoint real

number = 5
folder_name = "SUSTAINABILITY_115-5"
# Definir los tipos de dispositivos disponibles
device_types = [
    "switch", "router", "bridge", "repeater", "modern", "gateway", "firewall",
    "low_end_sensor", "high_end_sensor", "bulb", "energy_management", "lock",
    "security_alarm", "security_ip_camera", "appliance", "tv", "smartphone",
    "tablet", "pc", "smartwatch", "security_hub", "assistant_hub", "nas"
]

# Inicializar el diccionario new_config_devices con todos los tipos de dispositivos y valor 0
new_config_devices = {device: 0 for device in device_types}

# Distribuir aleatoriamente el número entre los dispositivos
remaining_number = number
while remaining_number > 0:
    # Elegir aleatoriamente un tipo de dispositivo
    device_type = random.choice(device_types)
    
    # Calcular el máximo que se puede asignar a este tipo de dispositivo
    max_allocation = min(remaining_number, 5 - new_config_devices[device_type])
    
    # Solo asignar si aún se puede asignar más a este tipo de dispositivo
    if max_allocation > 0:
        allocation = random.randint(1, max_allocation)  # Asignar al menos 1
        new_config_devices[device_type] += allocation
        remaining_number -= allocation

# Construir el payload completo
payload = {
    "new_config_devices": new_config_devices,
    "available_devices": {device: [] for device in device_types},
    "properties": {
        "security": "NONE",
        "usability": "NONE",
        "connectivity": "NONE",
        "sustainability": "VERYHIGH"
    }
}

# Convertir el payload a un string JSON con indentación bonita
json_payload = json.dumps(payload, indent=4)

# Definir los headers
headers = {
    'Content-Type': 'application/json'
}

# Crear la carpeta si no existe
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

# Archivos para almacenar los tiempos de respuesta y los nombres de los archivos
now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
request_filename = os.path.join(folder_name, f"request_{now}.json")
response_filename = os.path.join(folder_name, f"response_{now}.json")
csv_filename = os.path.join(folder_name, 'api_response_data.csv')
excel_filename = 'api_response_data.xlsx'

try:
    # Guardar la solicitud (request)
    with open(request_filename, 'w') as f_request:
        f_request.write(json_payload)

    # Medir el tiempo de respuesta
    start_time = time.time()
    response = requests.post(url, data=json.dumps(payload), headers=headers)
    end_time = time.time()
    response_time = end_time - start_time

    # Guardar la respuesta (response)
    response_json = response.json() if 'application/json' in response.headers.get('Content-Type', '') else response.text
    with open(response_filename, 'w') as f_response:
        json.dump(response_json, f_response, indent=4)

    # Imprimir la respuesta de manera formateada
    print('Status Code:', response.status_code)
    print('Solicitud guardada como:', request_filename)
    print('Respuesta guardada como:', response_filename)
    print('Tiempo de respuesta:', response_time, 'segundos')

    # Escribir los detalles en el archivo CSV
    with open(csv_filename, 'a', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        # Escribir la cabecera si el archivo está vacío
        if csvfile.tell() == 0:
            csv_writer.writerow(['Timestamp', 'Request Filename', 'Response Filename', 'Response Time (s)'])
        csv_writer.writerow([now, request_filename, response_filename, response_time])

    # Escribir los tiempos de respuesta en el archivo Excel
    if not os.path.exists(excel_filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'API Response Data'
        sheet.append([folder_name])  # Agrega la cabecera con el nombre de la carpeta
    else:
        workbook = load_workbook(excel_filename)
        sheet = workbook.active
    
    # Buscar la columna correspondiente al nombre de la carpeta
    col_idx = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == folder_name:
            col_idx = col
            break
    if col_idx is None:
        col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=col_idx).value = folder_name

    # Agregar el tiempo de respuesta en la columna correspondiente
    row_idx = 2  # Comenzamos a buscar desde la segunda fila
    while sheet.cell(row=row_idx, column=col_idx).value is not None:
        row_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = response_time

    workbook.save(excel_filename)

except Exception as e:
    print("Error al realizar la solicitud:", e)
